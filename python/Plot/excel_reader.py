from openpyxl import Workbook, load_workbook

class XLSXLoader():
    def __init__(self):
        self.wb = wb = load_workbook('plotdata.xlsx')
        self.ws = wb.active
        
        cols = list(self.ws.iter_cols())
        for col in cols:
            for cell in col:
                if cell.value == "DATA_X":
                    data_x = cell.column
                if cell.value == "DATA_Y":
                    data_y = cell.column
                """
                if cell.value == "DATA_X":
                    data_x = cell.column
                if cell.value == "DATA_Y":
                    data_y = cell.column
                """
        
        rows = list(self.ws.iter_rows())
        self.data_x, self.data_y = data_x, data_y
        self.max_data = len(rows) - 1
        
        self.xy_list = []
        for row in rows[1:]:
            self.xy_list.append([float(row[data_x-1].value), float(row[data_y-1].value)])
        print self.xy_list
    def getXYList(self):
        return self.xy_list
