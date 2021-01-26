from openpyxl import Workbook, load_workbook
from JimProgram import JimData
class XLSXLoader():
    def __init__(self):
        self.wb = wb = load_workbook('data.xlsx')
        self.ws = wb.active
        
        cols = list(self.ws.iter_cols())
        for col in cols:
            for cell in col:
                if cell.value == "TIME":
                    data_x = cell.column
                if cell.value == "TEMP":
                    data_y = cell.column
                
                if cell.value == "BPM":
                    data_b = cell.column
                if cell.value == "SPM":
                    data_s = cell.column
        
        rows = list(self.ws.iter_rows())
        self.data_x, self.data_y, self.data_b, self.data_s = data_x, data_y, data_b, data_s
        
        self.max_data = len(rows) - 1
        
        self.data = []
        for row in rows[1:]:
            self.data.append(JimData(int(row[data_y-1].value), int(row[data_b-1].value),
                                          int(row[data_s-1].value), int(row[data_x-1].value)))
        
    def getData(self): # NOt prefrrable, instead directly access self.data member..
        return self.data
        
    def refresh(self):
    
        data = []
        r = rows[1:]
        for row in r:
            data.append(JimData(int(row[data_y-1].value), int(row[data_b-1].value),
                                          int(row[data_s-1].value), int(row[data_x-1].value)))
        self.data = data
        
    def inputData(self, jimDataObj):
        self.data.append(jimDataObj)
        self.ws.append([jimDataObj.t, jimDataObj.k, jimDataObj.bpm, jimDataObj.spm])
    def save(self):
        self.wb.save('data.xlsx')
